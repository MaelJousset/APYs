
from seleniumbase import Driver
from selenium.webdriver.common.by import By

import os
import datetime

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

import pandas as pd
import matplotlib.pyplot as plt

DATAS_SHEET = "MySwap.xyz - Datas"
CHARTS_SHEET = "MySwap.xyz - Charts"


class MySwapPool:
    def __init__(self, name, tvl, vol24h, vol7d, apr7d):
        self.name = name
        self.tvl = tvl
        self.vol24h = vol24h
        self.vol7d = vol7d
        self.apr7d = apr7d


def init_myswap_driver():
    driver = Driver(uc=True, headless='headless')
    driver.get("https://www.myswap.xyz/#/charts")

    raw_datas = driver.find_elements(By.TAG_NAME, 'p')

    list_datas = []

    i = 0
    for elem in raw_datas:

        if elem.text == "#" and i == 0:
            index = raw_datas.index(elem)
            list_data = convert_myswap_datas(index, raw_datas)
            i += 1

            #write_xlsx(list_data)
            list_datas.append(list_data)

    driver.quit()

    return list_datas


def convert_myswap_datas(index, raw_datas):
    pool_list = []
    i = 0
    for elem in raw_datas[index:]:
        if elem.text.isdigit():
            cur_index = raw_datas.index(elem)

            p = MySwapPool(raw_datas[cur_index + 1].text,
                           raw_datas[cur_index + 2].text,
                           raw_datas[cur_index + 3].text,
                           raw_datas[cur_index + 4].text,
                           raw_datas[cur_index + 5].text)
            pool_list.append(p)
        elif elem.text == "#":
            i += 1
            if i == 2:
                break

    return pool_list


def write_myswap_xlsx(list_datas, filename):
    if not os.path.exists(filename):
        # Load the existing workbook
        workbook = Workbook()

        workbook.create_sheet(title=CHARTS_SHEET)
        myswap_datas_worksheet = workbook.create_sheet(title=DATAS_SHEET)
        workbook.remove(workbook['Sheet'])

        for datas in list_datas:
            my_swap_datas = generate_data_array(datas)

            for col_num, value in enumerate(my_swap_datas, start=1):
                cell = myswap_datas_worksheet.cell(row=1, column=col_num)
                cell.value = value

        workbook.save(filename)

    else:
        workbook = load_workbook(filename)
        myswap_datas_worksheet = workbook[DATAS_SHEET]
        myswap_charts_worksheet = workbook[CHARTS_SHEET]
        workbook.remove(myswap_charts_worksheet)
        workbook.create_sheet(title=CHARTS_SHEET)

        for datas in list_datas:
            my_swap_datas = generate_data_array(datas)

            next_row_number = myswap_datas_worksheet.max_row + 1
            for col_num, value in enumerate(my_swap_datas, start=1):
                cell = myswap_datas_worksheet.cell(row=next_row_number, column=col_num)
                cell.value = value

        workbook.save(filename)

    generate_myswap_charts(filename)


def generate_data_array(list_data):
    datas = [datetime.datetime.now().strftime('%d-%m-%Y'),
             datetime.datetime.now().strftime('%H:%M:%S')]

    for sel_pool in list_data:
        datas.append('')
        datas.append(sel_pool.name)
        datas.append(sel_pool.tvl)
        datas.append(sel_pool.vol24h)
        datas.append(sel_pool.vol7d)
        datas.append(sel_pool.apr7d)

    return datas


def generate_myswap_charts(filename):
    df = pd.read_excel(filename, sheet_name=DATAS_SHEET, header=None)
    df[0] = pd.to_datetime(df[0], format='%d-%m-%Y')

    df = df.dropna(axis=1, how='all')
    df = df.drop(df.columns[1], axis=1)

    for i in df.columns:
        if int(i) % 6 == 4 or int(i) % 6 == 5 or (int(i) % 6 == 0 and int(i) != 0):
            df = df.drop(i, axis=1)

    for i in df.columns:
        if i % 6 == 1:
            plt.figure()
            plt.plot(df[0], df[i])
            plt.xlabel('Date')
            plt.ylabel('APY %')
            plt.title(df[i-4][0])
            plt.xticks(rotation=45)
            plt.tight_layout()

            chart_image_file = 'images/chart_' + df[i-4][0].replace('/', '_') + '.png'
            plt.savefig(chart_image_file)

            workbook = load_workbook(filename)
            myswap_charts_worksheet = workbook[CHARTS_SHEET]

            # Insert the chart image into the worksheet
            img = Image(chart_image_file)

            if i % 12 == 1:
                col = 'B'
                row = str(i * 3 - 36)
            else:
                col = 'N'
                row = str(i * 3 - 18)

            myswap_charts_worksheet.add_image(img, col + row)

            workbook.save(filename)